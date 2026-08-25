"""
VoiceStamp Lead Finder — Streamlit web-app (snelle versie)
==========================================================
Nieuw t.o.v. de vorige versie:
  - VEEL SNELLER: websites worden parallel verwerkt (meerdere tegelijk).
  - MEER LEADS: bredere OpenStreetMap-categorieen per campagne (gratis, legaal).
  - Social links: Instagram / LinkedIn / Facebook per lead (handig voor je LinkedIn-aanpak).
  - OSM-cache: dezelfde provincie opnieuw ophalen gaat direct.
  - Live voortgang: je ziet meteen wat er gevonden wordt.

Lokaal: pip install streamlit requests dnspython openpyxl pandas ; streamlit run app.py
Online: zie LEES_MIJ_streamlit.md
"""

import re, io, time, csv, hashlib
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

import streamlit as st
import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import dns.resolver
    HAVE_DNS = True
except Exception:
    HAVE_DNS = False

# ------------------------------------------------------------------ CRM (Supabase, optioneel)
try:
    from supabase import create_client as _create_client
    SUPABASE_LIB = True
except Exception:
    SUPABASE_LIB = False

STATUS_OPTIES = ["Nieuw", "Gemaild", "Reactie", "Afspraak", "Klant", "Afgewezen", "Do not contact"]

def crm_beschikbaar():
    if not SUPABASE_LIB:
        return False
    try:
        return bool(st.secrets.get("SUPABASE_URL")) and bool(st.secrets.get("SUPABASE_ANON_KEY"))
    except Exception:
        return False

def crm_auth(email, password, registreren=False):
    client = _create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_ANON_KEY"])
    if registreren:
        res = client.auth.sign_up({"email": email, "password": password})
    else:
        res = client.auth.sign_in_with_password({"email": email, "password": password})
    sess = getattr(res, "session", None)
    if sess and getattr(sess, "access_token", None):
        try:
            client.postgrest.auth(sess.access_token)
        except Exception:
            pass
    return client, getattr(res, "user", None)

def crm_save(client, user_id, leads):
    bestaand = set()
    try:
        r = client.table("leads").select("email").eq("user_id", user_id).execute()
        bestaand = {row["email"] for row in (r.data or []) if row.get("email")}
    except Exception:
        pass
    rows = []
    for l in leads:
        e = l.get("email")
        if not e or e in bestaand:
            continue
        bestaand.add(e)
        rows.append({"user_id": user_id, "company_name": l["naam"], "segment": l["type"],
                     "website": l.get("website", ""), "email": e, "phone": l.get("telefoon", ""),
                     "score": l.get("_score"), "hook": l.get("haak", ""),
                     "channel_advice": l.get("_kanaal", ""), "status": "Nieuw"})
    if rows:
        client.table("leads").insert(rows).execute()
    return len(rows)

def crm_load(client, user_id):
    r = client.table("leads").select("*").eq("user_id", user_id).order("score", desc=True).execute()
    return r.data or []

def crm_update(client, row_id, user_id, fields):
    # Defensief: filter op id EN user_id (bovenop row-level security in Supabase).
    client.table("leads").update(fields).eq("id", row_id).eq("user_id", user_id).execute()

def crm_delete(client, row_id, user_id):
    client.table("leads").delete().eq("id", row_id).eq("user_id", user_id).execute()

def crm_delete_all(client, user_id):
    r = client.table("leads").delete().eq("user_id", user_id).execute()
    return len(r.data or [])

def crm_add_activity(client, user_id, lead_id, type_, description=""):
    try:
        client.table("activities").insert({
            "user_id": user_id, "lead_id": lead_id, "type": type_, "description": description
        }).execute()
    except Exception:
        pass

def crm_load_activities(client, user_id, lead_id):
    try:
        r = (client.table("activities").select("*")
             .eq("user_id", user_id).eq("lead_id", lead_id)
             .order("created_at", desc=True).execute())
        return r.data or []
    except Exception:
        return []

def crm_lead_id_by_email(client, user_id, email):
    try:
        r = client.table("leads").select("id").eq("user_id", user_id).eq("email", email).limit(1).execute()
        if r.data:
            return r.data[0]["id"]
    except Exception:
        pass
    return None

def crm_sent_today(client, user_id):
    """Aantal 'Mail verstuurd'-activiteiten van vandaag (voor de dagelijkse limiet)."""
    from datetime import date
    try:
        r = (client.table("activities").select("id")
             .eq("user_id", user_id).eq("type", "Mail verstuurd")
             .gte("created_at", date.today().isoformat()).execute())
        return len(r.data or [])
    except Exception:
        return 0

# ---------------------------------------------------------------- Zoho-mail (SMTP)
import smtplib, ssl
from email.mime.text import MIMEText
from email.utils import formataddr

MAIL_DAG_LIMIET = 10

def zoho_send(host, port, user, app_pw, from_name, to_addr, subject, body):
    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = formataddr((from_name, user))
    msg["To"] = to_addr
    ctx = ssl.create_default_context()
    with smtplib.SMTP_SSL(host, int(port), context=ctx, timeout=20) as s:
        s.login(user, app_pw)
        s.sendmail(user, [to_addr], msg.as_string())

# Koppel Nederlandse (of Engelse) kolomnamen aan de CRM-velden.
IMPORT_MAP = {
    "company_name": ["naam", "bedrijf", "company_name", "company", "bedrijfsnaam"],
    "email": ["e-mail", "email", "e-mail (beste ingang)", "mail", "e-mailadres"],
    "segment": ["type", "segment", "categorie"],
    "website": ["website", "url", "site"],
    "phone": ["telefoon", "phone", "tel", "telefoonnummer"],
    "status": ["status"],
    "note": ["notities", "notitie", "note", "opmerkingen"],
}

def _kies_kolom(kolommen_laag, opties):
    for o in opties:
        if o in kolommen_laag:
            return kolommen_laag[o]
    return None

def crm_import_uit_bytes(client, user_id, data_bytes, filename):
    """Lees een geuploade Excel/CSV en zet de leads in het CRM. Geeft (toegevoegd, overgeslagen)."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        df = pd.read_excel(io.BytesIO(data_bytes))
    else:
        df = pd.read_csv(io.BytesIO(data_bytes))
    # kolomnamen -> {kleine letter: originele naam}
    kol = {str(c).strip().lower(): c for c in df.columns}
    mapping = {veld: _kies_kolom(kol, opties) for veld, opties in IMPORT_MAP.items()}
    if not mapping.get("email"):
        raise ValueError("Geen e-mailkolom gevonden. Zorg dat er een kolom 'E-mail' in staat.")
    # bestaande e-mails ophalen om dubbel te voorkomen
    bestaand = set()
    try:
        r = client.table("leads").select("email").eq("user_id", user_id).execute()
        bestaand = {row["email"] for row in (r.data or []) if row.get("email")}
    except Exception:
        pass
    rows, overgeslagen = [], 0
    for _, rij in df.iterrows():
        email = str(rij.get(mapping["email"], "") or "").strip().lower()
        if not email or "@" not in email or email in bestaand:
            overgeslagen += 1
            continue
        bestaand.add(email)
        def val(veld):
            c = mapping.get(veld)
            if not c: return ""
            v = rij.get(c, "")
            return "" if pd.isna(v) else str(v).strip()
        rows.append({"user_id": user_id, "company_name": val("company_name"), "email": email,
                     "segment": val("segment"), "website": val("website"), "phone": val("phone"),
                     "status": val("status") or "Nieuw", "note": val("note")})
    # in blokjes wegschrijven
    for i in range(0, len(rows), 100):
        client.table("leads").insert(rows[i:i+100]).execute()
    return len(rows), overgeslagen


# ------------------------------------------------------------------ constants
OVERPASS = ["https://overpass-api.de/api/interpreter",
            "https://overpass.kumi.systems/api/interpreter",
            "https://maps.mail.ru/osm/tools/overpass/api/interpreter",
            "https://overpass.openstreetmap.ru/api/interpreter",
            "https://overpass.private.coffee/api/interpreter"]
UA = "VoiceStamp-LeadFinder/streamlit (zakelijk gebruik; contact via voicestamp.nl)"
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
BAD = ("example.", "sentry.", "wixpress.", ".png", ".jpg", ".gif", "@2x",
       "your-email", "email@", "name@", "domain.com", "@sentry")
# Kortere padlijst = sneller. Homepage + de paar meest waarschijnlijke contactpaginas.
PATHS = ["", "contact", "over-ons", "info", "reserveren", "about"]
ROLE = ("info", "contact", "welkom", "boeking", "boekingen", "reserveringen",
        "reservering", "hallo", "mail", "receptie", "office", "sales")
KETEN = ("hotels", "resorts", "group", "vakantieparken", "landal", "roompot",
         "rcn", "huttopia", "fletcher", "ardoer", "europarcs", "oostappen")
SOCIAL_RE = {
    "instagram": re.compile(r"https?://(?:www\.)?instagram\.com/[A-Za-z0-9_.]+"),
    "linkedin": re.compile(r"https?://(?:[a-z]{2,3}\.)?linkedin\.com/(?:company|in)/[A-Za-z0-9_%\-]+"),
    "facebook": re.compile(r"https?://(?:www\.)?facebook\.com/[A-Za-z0-9_.\-]+"),
}

# --- Netjes blijven tegenover de bronnen: rate limiting + robots + retries ---
import threading
from urllib import robotparser

_domain_lock = threading.Lock()
_last_hit = {}          # domein -> laatste tijdstip
_robots_cache = {}      # domein -> RobotFileParser (of None)
MIN_INTERVAL = 1.0      # minimaal 1 seconde tussen verzoeken naar hetzelfde domein

def _respect_delay(domain):
    """Wacht indien nodig zodat we niet te snel achter elkaar hetzelfde domein raken."""
    with _domain_lock:
        now = time.time()
        wait = MIN_INTERVAL - (now - _last_hit.get(domain, 0))
        if wait > 0:
            time.sleep(wait)
        _last_hit[domain] = time.time()

def _robots_ok(base, path):
    """Controleer robots.txt; bij twijfel toestaan (maar netjes blijven qua tempo)."""
    domain = urlparse(base).netloc
    rp = _robots_cache.get(domain, "missing")
    if rp == "missing":
        rp = robotparser.RobotFileParser()
        try:
            rp.set_url(f"{urlparse(base).scheme}://{domain}/robots.txt")
            rp.read()
        except Exception:
            rp = None
        _robots_cache[domain] = rp
    if rp is None:
        return True
    try:
        return rp.can_fetch(UA, f"{base}/{path}".rstrip("/"))
    except Exception:
        return True

def polite_get(url, domain, timeout=8, retries=2):
    """GET met rate limiting en een enkele nette retry bij een tijdelijke fout."""
    for attempt in range(retries):
        _respect_delay(domain)
        try:
            r = requests.get(url, headers={"User-Agent": UA}, timeout=timeout)
            if r.status_code in (429, 503) and attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
                continue
            return r
        except Exception:
            if attempt < retries - 1:
                time.sleep(1)
                continue
            raise
    return None
PROVINCIES = ["Groningen", "Fryslân", "Drenthe", "Overijssel", "Flevoland",
              "Gelderland", "Utrecht", "Noord-Holland", "Zuid-Holland",
              "Zeeland", "Noord-Brabant", "Limburg"]

# Bredere presets = meer leads. Elke osm-value krijgt (label, template-segment).
PRESETS = {
    "Verblijf (campings, hotels, B&B's)": {
        "typemap": {
            "camp_site": ("Camping", "natuurcamping"), "caravan_site": ("Camperplaats", "natuurcamping"),
            "guest_house": ("B&B / guest house", "bnb"), "hotel": ("Hotel", "hotel"),
            "hostel": ("Hostel", "hotel"), "chalet": ("Chalet/huisjes", "natuurcamping"),
            "motel": ("Motel", "hotel"), "apartment": ("Vakantieappartement", "bnb"),
            "alpine_hut": ("Berghut/verblijf", "natuurcamping"),
        },
        "keys": {"tourism": ["camp_site", "caravan_site", "guest_house", "hotel",
                             "hostel", "chalet", "motel", "apartment", "alpine_hut"]},
    },
    "Streek (wijn, kaas, brouwers, boerderijwinkels)": {
        "typemap": {
            "winery": ("Wijngaard", "streek"), "brewery": ("Brouwerij", "streek"),
            "distillery": ("Distilleerderij", "streek"), "cheese_making": ("Kaasmakerij", "streek"),
            "wine": ("Wijnhandel", "streek"), "cheese": ("Kaaswinkel", "streek"),
            "farm": ("Boerderijwinkel", "streek"), "dairy": ("Zuivel/boerderij", "streek"),
            "greengrocer": ("Groente/streekwinkel", "streek"), "deli": ("Delicatessen", "streek"),
            "bakery": ("Bakkerij", "streek"), "butcher": ("Slagerij", "streek"),
            "confectionery": ("Chocolaterie/snoep", "streek"), "chocolate": ("Chocolatier", "streek"),
            "honey": ("Imker/honing", "streek"),
        },
        "keys": {"craft": ["winery", "brewery", "distillery", "cheese_making"],
                 "shop": ["wine", "cheese", "farm", "dairy", "greengrocer", "deli",
                          "bakery", "butcher", "confectionery", "chocolate", "honey"]},
    },
    "Attracties (dierentuinen, kinderboerderijen, parken)": {
        "typemap": {
            "zoo": ("Dierentuin/kinderboerderij", "attractie"),
            "theme_park": ("Attractiepark", "attractie"), "aquarium": ("Aquarium", "attractie"),
            "water_park": ("Waterpark", "attractie"),
        },
        "keys": {"tourism": ["zoo", "theme_park", "aquarium"], "leisure": ["water_park"]},
    },
    "Erfgoed (musea, kastelen, landgoederen)": {
        "typemap": {
            "museum": ("Museum", "erfgoed"), "gallery": ("Galerie", "erfgoed"),
            "castle": ("Kasteel", "erfgoed"), "manor": ("Landgoed", "erfgoed"),
            "fort": ("Fort", "erfgoed"), "monastery": ("Klooster", "erfgoed"),
        },
        "keys": {"tourism": ["museum", "gallery"],
                 "historic": ["castle", "manor", "fort", "monastery"]},
    },
    "Horeca (restaurants, cafés)": {
        "typemap": {
            "restaurant": ("Restaurant", "horeca"), "cafe": ("Café/lunchroom", "horeca"),
            "pub": ("Eetcafé/pub", "horeca"), "bistro": ("Bistro", "horeca"),
            "ice_cream": ("IJssalon", "horeca"),
        },
        "keys": {"amenity": ["restaurant", "cafe", "pub", "bistro", "ice_cream"]},
    },
    "Winkels met een verhaal (cadeau, kunst, sieraden)": {
        "typemap": {
            "gift": ("Cadeauwinkel", "winkel"), "art": ("Kunstwinkel/galerie", "winkel"),
            "books": ("Boekwinkel", "winkel"), "jewelry": ("Sieradenwinkel", "winkel"),
            "antiques": ("Antiek", "winkel"), "interior_decoration": ("Interieurwinkel", "winkel"),
            "pottery": ("Keramiek/aardewerk", "winkel"), "craft": ("Hobby/ambacht", "winkel"),
            "musical_instrument": ("Muziekwinkel", "winkel"), "watches": ("Horlogewinkel", "winkel"),
        },
        "keys": {"shop": ["gift", "art", "books", "jewelry", "antiques",
                          "interior_decoration", "pottery", "craft",
                          "musical_instrument", "watches"]},
    },
    "Bloemisten": {
        "typemap": {"florist": ("Bloemist", "bloemist"), "garden_centre": ("Tuincentrum", "bloemist")},
        "keys": {"shop": ["florist", "garden_centre"]},
    },
}

# Omgekeerde koppeling: opgeslagen type-label (bijv. "Camping") terug naar segment ("natuurcamping").
LABEL_TO_SEG = {}
for _p in PRESETS.values():
    for _osmval, (_label, _seg) in _p["typemap"].items():
        LABEL_TO_SEG[_label] = _seg

LINKS = ("\u2022 Website: www.voicestamp.nl\n"
         "\u2022 Hoe het werkt: www.voicestamp.nl/how-it-works\n"
         "\u2022 Instagram: @voicestamp.nl\n"
         "\u2022 LinkedIn: VoiceStamp")
SIG = ("Met vriendelijke groet,\nYusuf Tatlicioglu\n0646756497\n\n"
       "Founder VoiceStamp\nhttps://www.voicestamp.nl/")
APP_LINE = ("Ik zag dat jullie al met een app werken, dus voor de duidelijkheid: VoiceStamp is "
            "g\u00e9\u00e9n tweede app. Een gast hoeft niets te downloaden en geen account te maken. "
            "Ik zie het naast jullie app staan, niet in de weg.")

def _mid_placeholder():
    pass

MID = {
 "natuurcamping": ("Wat als een gast dat verhaal niet alleen leest, maar het ook rechtstreeks van "
   "jullie hoort?\n\nMet VoiceStamp scan je een eenvoudige stempel (VoiceStamp) op een plek, en "
   "opent direct een persoonlijke audioboodschap met een landingspagina. Geen app, geen account: "
   "juist minder drukte, niet meer. Een warm welkom, een verhaal over de omgeving, of een wandeltip."),
 "bnb": ("Wat als een gast dat verhaal ook van jullie hoort, op het moment dat hij binnenkomt?\n\n"
   "Met VoiceStamp scan je een eenvoudige stempel (VoiceStamp) op de kamer, en opent direct een "
   "persoonlijke audioboodschap met een landingspagina. Geen app, geen account. Een warm welkom in "
   "je eigen stem, het verhaal van het pand, of een tip voor de omgeving."),
 "hotel": ("Wat als een gast dat verhaal hoort op het moment dat hij binnenkomt?\n\nMet VoiceStamp "
   "scan je een eenvoudige stempel (VoiceStamp) op de kamer of in de lobby, en opent direct een "
   "audioboodschap met een landingspagina. Geen app, geen account. Een warm welkom, het verhaal van "
   "de plek, of info die nu aan de balie wordt gevraagd."),
 "streek": ("Want dat verhaal vertel je in de winkel, maar het gaat niet mee met het product dat "
   "iemand koopt.\n\nMet VoiceStamp scan je een eenvoudige stempel (VoiceStamp) op de verpakking, en "
   "opent direct een audioboodschap met een landingspagina. Geen app, geen account. Waarom een "
   "ingredi\u00ebnt is gekozen, hoe iets gemaakt wordt, of een welkom bij een proeverij."),
 "attractie": ("Wat als een bezoeker bij een dier of attractie een verzorger het hoort vertellen, in "
   "plaats van een bordje te lezen?\n\nMet VoiceStamp scan je een eenvoudige stempel (VoiceStamp) bij "
   "een verblijf, en opent direct een audioboodschap met een landingspagina. Geen app \u2014 en het "
   "werkt ook voor kinderen die nog niet lezen. Het verhaal van een dier, een welkom, of het "
   "dagprogramma."),
 "erfgoed": ("Wat als een bezoeker het verhaal hoort in de stem van een gids of conservator?\n\nMet "
   "VoiceStamp scan je een eenvoudige stempel (VoiceStamp) bij een object of monument, en opent "
   "direct een audioboodschap met een landingspagina. Geen app, geen account. Zo maak je het verhaal "
   "toegankelijk, ook buiten de rondleidingen om."),
 "horeca": ("Wat als een gast niet alleen leest wat er op de kaart staat, maar de chef zelf hoort "
   "vertellen waarom een gerecht er staat?\n\nMet VoiceStamp scan je een eenvoudige stempel "
   "(VoiceStamp) op de menukaart of op tafel, en opent direct een audioboodschap met een "
   "landingspagina. Geen app, geen account. Een welkom bij binnenkomst, het verhaal achter een "
   "streekgerecht of leverancier, of een tip voor een volgende keer."),
 "winkel": ("Want dat verhaal vertel je in de winkel, maar het gaat niet mee met het product dat "
   "iemand koopt.\n\nMet VoiceStamp scan je een eenvoudige stempel (VoiceStamp) op het product of de "
   "verpakking, en opent direct een audioboodschap met een landingspagina. Geen app, geen account. "
   "Waarom een stuk gemaakt is, wie de maker is, of hoe je het het beste gebruikt of onderhoudt."),
 "bloemist": ("Wat als een boeket zelf een boodschap kon meegeven, in de stem van de afzender?\n\n"
   "Met VoiceStamp scan je een eenvoudige stempel (VoiceStamp) bij het boeket, in plaats van het "
   "kleine kaartje, en opent direct een persoonlijke audioboodschap met een landingspagina. Geen "
   "app, geen account. Juist rond momenten als Moederdag of Valentijn maakt een stem het verschil."),
}
SUBJECTS = {
 "natuurcamping": ["Wat als jullie plek zelf haar verhaal kon vertellen?", "Een stem bij jullie plek, zonder app"],
 "bnb": ["Wat als jullie plek zelf haar verhaal kon vertellen?", "Een persoonlijk welkom op de kamer"],
 "hotel": ["Wat als jullie plek zelf haar verhaal kon vertellen?", "Minder vragen aan de balie"],
 "streek": ["Wat als jullie product zelf zijn verhaal kon vertellen?", "Het verhaal achter het product"],
 "attractie": ["Wat als jullie dieren hun eigen verhaal konden vertellen?", "Een stem bij elk verblijf"],
 "erfgoed": ["Wat als jullie collectie zelf haar verhaal kon vertellen?", "Een stem bij het monument"],
 "horeca": ["Wat als jullie gerechten hun eigen verhaal konden vertellen?", "Een woord van de chef, op tafel"],
 "winkel": ["Wat als jullie producten zelf hun verhaal konden vertellen?", "Het verhaal achter het product, bij de klant thuis"],
 "bloemist": ["Wat als een boeket zelf iets kon zeggen?", "Een stem bij de bloemen, geen kaartje"],
}
FOLLOWUP_2 = ("Beste {aanhef},\n\nKorte opvolging op mijn vorige bericht. Ik snap dat het druk is, "
              "dus kort: met {merk} horen jullie mensen een stem op de plek zelf, in plaats van een "
              "bordje te lezen. Geen app, geen account.\n\nZal ik het een keer laten zien?\n\n"
              "Met vriendelijke groet,")
FOLLOWUP_3 = ("Beste {aanhef},\n\nLaatste bericht van mijn kant, daarna laat ik het rusten. Mocht je "
              "ooit willen dat jullie verhaal ook echt te h\u00f3ren is op de plek zelf, dan weet je me "
              "te vinden.\n\nMet vriendelijke groet,")

# Prikkelende onderwerpregels per segment (in plaats van een 'wat als'-vraag).
SUBJECTS_STD = {
    "natuurcamping": "Wat als gasten jullie camping ook konden horen?",
    "bnb": "Wat als gasten jullie plek ook konden horen?",
    "hotel": "Wat als gasten jullie hotel ook konden horen?",
    "streek": "Wat als jullie producten iets konden vertellen?",
    "winkel": "Wat als jullie producten iets konden vertellen?",
    "horeca": "Wat als jullie gerechten iets konden vertellen?",
    "bloemist": "Wat als een boeket iets kon zeggen?",
    "attractie": "Wat als jullie dieren iets konden vertellen?",
    "erfgoed": "Wat als jullie plek ook te horen was?",
}

# Kerntekst per segment (zonder aanhef, zonder links, zonder groet).
STD = {
 "natuurcamping": {"segwoord": "camping", "kern":
   "Wat als gasten jullie camping niet alleen zien, maar ook kunnen horen?\n\n"
   "Een welkom bij aankomst. Het verhaal achter de plek. Een persoonlijke tip voor de mooiste "
   "wandeling.\n\n"
   "Met VoiceStamp scant een gast een eenvoudige stempel en hoort direct jullie boodschap, op het "
   "moment en de plek waar die relevant is.\n\n"
   "Geen app. Geen account. Gewoon luisteren.\n\n"
   "Een VoiceStamp kan staan op de kampeerplek, bij de receptie of aan het begin van een route. In "
   "jullie eigen stem.\n\n"
   "En dat is maar \u00e9\u00e9n voorbeeld. VoiceStamp kan overal op de camping een verhaal, boodschap of "
   "moment een stem geven.\n\n"
   "Zo geef je gasten iets mee dat op een bordje vaak verloren gaat, zonder extra drukte voor je "
   "team."},
 "bnb": {"segwoord": "B&B", "kern":
   "Wat als gasten jullie plek niet alleen zien, maar ook kunnen horen?\n\n"
   "Een welkom bij aankomst. Het verhaal van het pand. Een persoonlijke tip voor de omgeving.\n\n"
   "Met VoiceStamp scant een gast een eenvoudige stempel en hoort direct jullie boodschap, op het "
   "moment en de plek waar die relevant is.\n\n"
   "Geen app. Geen account. Gewoon luisteren.\n\n"
   "Een VoiceStamp kan op de kamer, in de gezamenlijke ruimte of bij een bijzonder detail van het "
   "pand. In jullie eigen stem.\n\n"
   "En dat is maar \u00e9\u00e9n voorbeeld. VoiceStamp kan overal in jullie B&B een verhaal, boodschap of "
   "moment een stem geven.\n\n"
   "Zo geef je gasten iets mee dat op een briefje vaak verloren gaat, zonder extra drukte voor je "
   "team."},
 "hotel": {"segwoord": "hotel", "kern":
   "Wat als gasten jullie hotel niet alleen zien, maar ook kunnen horen?\n\n"
   "Een welkom bij aankomst. Het verhaal van het pand. Een tip voor de omgeving.\n\n"
   "Met VoiceStamp scant een gast een eenvoudige stempel en hoort direct jullie boodschap, op het "
   "moment en de plek waar die relevant is.\n\n"
   "Geen app. Geen account. Gewoon luisteren.\n\n"
   "Een VoiceStamp kan op de kamer, in de lobby of in het restaurant. In jullie eigen stem.\n\n"
   "En dat is maar \u00e9\u00e9n voorbeeld. VoiceStamp kan overal in jullie hotel een verhaal, boodschap of "
   "moment een stem geven.\n\n"
   "Zo geef je gasten iets mee dat op een gelamineerd vel vaak verloren gaat, en vang je meteen "
   "vragen af die anders bij de receptie komen."},
 "streek": {"segwoord": "zaak", "kern":
   "Wat als jullie producten niet alleen gezien worden, maar ook iets kunnen vertellen?\n\n"
   "Waarom een druif, een kaas of een ingredi\u00ebnt gekozen is. Hoe iets gemaakt wordt. Een woord "
   "van de maker zelf.\n\n"
   "Met VoiceStamp scant een klant een eenvoudige stempel en hoort direct jullie boodschap, op het "
   "moment en de plek waar die relevant is.\n\n"
   "Geen app. Geen account. Gewoon luisteren.\n\n"
   "Een VoiceStamp kan op de verpakking, bij een schap of aan de proeftafel. In jullie eigen stem.\n\n"
   "En dat is maar \u00e9\u00e9n voorbeeld. VoiceStamp kan bij elk product een verhaal, boodschap of moment "
   "een stem geven.\n\n"
   "Zo geef je klanten iets mee dat op een etiket vaak verloren gaat, ook nog thuis aan tafel."},
 "winkel": {"segwoord": "winkel", "kern":
   "Wat als jullie producten niet alleen gezien worden, maar ook iets kunnen vertellen?\n\n"
   "Wie het maakte. Waarom het bijzonder is. Hoe je het het beste gebruikt.\n\n"
   "Met VoiceStamp scant een klant een eenvoudige stempel en hoort direct jullie boodschap, op het "
   "moment en de plek waar die relevant is.\n\n"
   "Geen app. Geen account. Gewoon luisteren.\n\n"
   "Een VoiceStamp kan op het product, op de verpakking of bij een schap. In de stem van de maker.\n\n"
   "En dat is maar \u00e9\u00e9n voorbeeld. VoiceStamp kan bij elk product een verhaal, boodschap of moment "
   "een stem geven.\n\n"
   "Zo geef je klanten iets mee dat in de winkel verteld wordt, maar anders niet meegaat naar huis."},
 "horeca": {"segwoord": "zaak", "kern":
   "Wat als gasten jullie gerechten niet alleen proeven, maar er ook iets over kunnen horen?\n\n"
   "Waarom een gerecht op de kaart staat. Waar een product vandaan komt. Een woord van de chef.\n\n"
   "Met VoiceStamp scant een gast een eenvoudige stempel en hoort direct jullie boodschap, op het "
   "moment en de plek waar die relevant is.\n\n"
   "Geen app. Geen account. Gewoon luisteren.\n\n"
   "Een VoiceStamp kan op de menukaart, op tafel of bij binnenkomst. In jullie eigen stem.\n\n"
   "En dat is maar \u00e9\u00e9n voorbeeld. VoiceStamp kan overal in jullie zaak een verhaal, boodschap of "
   "moment een stem geven.\n\n"
   "Zo geef je gasten iets mee dat op een drukke avond vaak verloren gaat, zonder extra werk voor de "
   "bediening."},
 "bloemist": {"segwoord": "winkel", "kern":
   "Wat als een boeket niet alleen gezien wordt, maar ook iets kan zeggen?\n\n"
   "Een persoonlijke felicitatie. Een woord van troost. Gewoon: ik denk aan je.\n\n"
   "Met VoiceStamp scant de ontvanger een eenvoudige stempel en hoort direct de boodschap, op het "
   "moment dat het telt.\n\n"
   "Geen app. Geen account. Gewoon luisteren.\n\n"
   "Een VoiceStamp kan bij het boeket, in plaats van het kleine kaartje. In de eigen stem van de "
   "afzender.\n\n"
   "En dat is maar \u00e9\u00e9n voorbeeld. VoiceStamp kan bij elk boeket een boodschap of moment een stem "
   "geven.\n\n"
   "Juist rond Moederdag en Valentijn maakt een stem het verschil, zonder extra gedoe voor jullie."},
 "attractie": {"segwoord": "park", "kern":
   "Wat als bezoekers jullie dieren en attracties niet alleen zien, maar er ook iets over kunnen "
   "horen?\n\n"
   "Het verhaal van een dier. Een welkom bij binnenkomst. Het programma van die dag.\n\n"
   "Met VoiceStamp scant een bezoeker een eenvoudige stempel en hoort direct jullie boodschap, op "
   "het moment en de plek waar die relevant is.\n\n"
   "Geen app. Geen account. Gewoon luisteren.\n\n"
   "Een VoiceStamp kan bij een verblijf, langs de route of bij de ingang. Verteld door een "
   "verzorger.\n\n"
   "En dat is maar \u00e9\u00e9n voorbeeld. VoiceStamp kan overal in jullie park een verhaal, boodschap of "
   "moment een stem geven.\n\n"
   "Zo geef je bezoekers iets mee dat op een bordje vaak verloren gaat, en dat ook werkt voor "
   "kinderen die nog niet lezen."},
 "erfgoed": {"segwoord": "organisatie", "kern":
   "Wat als bezoekers jullie plek niet alleen zien, maar ook kunnen horen?\n\n"
   "Het verhaal achter een stuk. Een route langs de hoogtepunten. Een welkom bij binnenkomst.\n\n"
   "Met VoiceStamp scant een bezoeker een eenvoudige stempel en hoort direct jullie boodschap, op "
   "het moment en de plek waar die relevant is.\n\n"
   "Geen app. Geen account. Gewoon luisteren.\n\n"
   "Een VoiceStamp kan bij een object, een monument of langs een route. In de stem van een gids of "
   "conservator.\n\n"
   "En dat is maar \u00e9\u00e9n voorbeeld. VoiceStamp kan overal in jullie collectie een verhaal, boodschap "
   "of moment een stem geven.\n\n"
   "Zo maak je jullie verhaal toegankelijk, ook buiten de rondleidingen om."},
}

DEFAULT_OPENER = {
 "natuurcamping": "Tijdens het bekijken van jullie website viel me op hoeveel aandacht jullie besteden aan rust en natuur.",
 "bnb": "Tijdens het bekijken van jullie website viel me op hoe persoonlijk jullie gasten ontvangen.",
 "hotel": "Tijdens het bekijken van jullie website viel me op hoeveel karakter jullie plek heeft.",
 "streek": "Tijdens het bekijken van jullie website viel me op hoeveel verhaal er in jullie producten zit.",
 "attractie": "Tijdens het bekijken van jullie website viel me op hoeveel er bij jullie te beleven is.",
 "erfgoed": "Tijdens het bekijken van jullie website viel me op hoeveel verhaal jullie plek draagt.",
 "horeca": "Tijdens het bekijken van jullie website viel me op hoeveel aandacht jullie besteden aan de gerechten en de sfeer.",
 "winkel": "Tijdens het bekijken van jullie website viel me op hoeveel verhaal er in jullie producten zit.",
 "bloemist": "Tijdens het bekijken van jullie website viel me op hoeveel gevoel jullie in jullie boeketten leggen.",
}
KEYWORDS = ("rust", "natuur", "gastvrij", "persoonlijk", "familie", "verhaal", "duurzaam",
            "welkom", "beleef", "genieten", "monument", "ambacht", "streek", "traditie",
            "avontuur", "kinderen", "dieren")

# --- Haak-vinder: signalen die een reden geven om contact op te nemen ---
# label -> regex (op kleine letters). De eerste match per lead wordt de "haak".
SIGNALS = {
    "Prijs/onderscheiding": re.compile(
        r"\b(award|bekroond|genomineerd|winnaar|onderscheiding|pincamp|green key|"
        r"michelin|beste [a-z]+ van|verkozen tot)\b"),
    "Lange historie": re.compile(
        r"\b(sinds (1[6-9]\d\d|20[01]\d)|opgericht in|al \d{2,3} jaar|"
        r"\d+e generatie|generatie op generatie|eeuwenoud)\b"),
    "Familiebedrijf": re.compile(r"\b(familiebedrijf|van vader op zoon|onze familie|familie [a-z]+ runt)\b"),
    "Duurzaam / B Corp": re.compile(r"\b(duurzaam|biologisch|b corp|co2-neutraal|klimaatneutraal|permacultuur)\b"),
    "Net geopend / nieuw": re.compile(r"\b(net geopend|sinds kort|nieuw seizoen|onlangs geopend|kersvers)\b"),
    "Eigen verhaal/oprichter": re.compile(r"\b(ons verhaal|het verhaal achter|de oprichter|opgericht door)\b"),
    "Ambachtelijk / eigen productie": re.compile(r"\b(ambachtelijk|zelfgemaakt|eigen productie|met de hand gemaakt|hoeve-eigen)\b"),
}
# Als een van deze matcht heeft het bedrijf al een app -> pas de pitch aan.
APP_HINTS = re.compile(r"(download (?:onze|de) app|in de app store|google play|via onze app|onze eigen app)")
SIGNAL_WORDS = ("award", "bekroond", "genomineerd", "winnaar", "michelin", "green key",
                "familiebedrijf", "generatie", "sinds 1", "sinds 20", "ambachtelijk",
                "duurzaam", "oprichter", "ons verhaal", "b corp")

def detect_signals(text):
    """Geef (haak-label of '', heeft_app-bool) terug op basis van de sitetekst."""
    tl = text.lower()
    haak = ""
    for label, rx in SIGNALS.items():
        if rx.search(tl):
            haak = label
            break
    return haak, bool(APP_HINTS.search(tl))

# ------------------------------------------------------------------ OSM
def q_build(area, keys, whole):
    parts = []
    for k, vals in keys.items():
        for v in vals:
            parts.append(f'node["{k}"="{v}"](area.a);way["{k}"="{v}"](area.a);'
                         f'relation["{k}"="{v}"](area.a);')
    ab = ('area["ISO3166-1"="NL"][admin_level=2]->.a;' if whole
          else f'area["name"="{area}"]["admin_level"="4"]->.a;')
    return f'[out:json][timeout:180];{ab}({"".join(parts)});out center tags;'

@st.cache_data(ttl=7 * 24 * 3600, show_spinner=False)
def osm_cached(area, keys_items, whole):
    keys = {k: list(v) for k, v in keys_items}
    q = q_build(area, keys, whole)
    laatste = None
    # 2 rondes over alle servers, met een korte pauze, zodat tijdelijke 500/502-fouten worden opgevangen
    for ronde in range(2):
        for ep in OVERPASS:
            try:
                r = requests.post(ep, data={"data": q},
                                  headers={"User-Agent": UA}, timeout=180)
                if r.status_code in (429, 500, 502, 503, 504):
                    laatste = f"{r.status_code} van {urlparse(ep).netloc}"
                    time.sleep(1.5)
                    continue
                r.raise_for_status()
                els = r.json().get("elements", [])
                return els
            except Exception as e:
                laatste = f"{type(e).__name__} bij {urlparse(ep).netloc}"
                time.sleep(1.0)
                continue
        time.sleep(2)
    raise RuntimeError(
        "De OpenStreetMap-servers reageren nu niet (dit ligt aan hun kant, niet aan de app). "
        f"Laatste melding: {laatste}. Probeer het over een paar minuten opnieuw, of kies "
        "eventueel een kleinere provincie.")

def parse(els, typemap):
    out = []
    for el in els:
        t = el.get("tags", {}); name = t.get("name")
        if not name: continue
        osmval = next((t.get(k) for k in ("tourism", "craft", "shop", "historic", "leisure", "amenity")
                       if t.get(k) in typemap), None)
        if not osmval: continue
        label, seg = typemap[osmval]
        email = (t.get("contact:email") or t.get("email") or "").strip().lower()
        out.append({"naam": name, "type": label, "seg": seg,
                    "plaats": (t.get("addr:city") or t.get("addr:place") or "").strip(),
                    "email": email if EMAIL_RE.fullmatch(email or "") else "",
                    "website": (t.get("contact:website") or t.get("website") or "").strip(),
                    "telefoon": (t.get("contact:phone") or t.get("phone") or "").strip(),
                    "instagram": (t.get("contact:instagram") or ""),
                    "linkedin": "", "facebook": (t.get("contact:facebook") or ""),
                    "haak": "", "heeft_app": False,
                    "bron": "OpenStreetMap", "opener": ""})
    return out

# ------------------------------------------------------------------ website
def norm(u):
    if not u: return ""
    if not u.startswith(("http://", "https://")): u = "https://" + u
    return u.rstrip("/")

def detag(h):
    h = re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", h, flags=re.S | re.I)
    return re.sub(r"<[^>]+>", " ", h)

# Vaste, altijd-correcte openingszinnen per gevonden haak (geen zinsontleding nodig).
HOOK_OPENERS = {
    "Prijs/onderscheiding": "Tijdens het bekijken van jullie website viel me op dat jullie in de prijzen zijn gevallen.",
    "Lange historie": "Tijdens het bekijken van jullie website viel me op hoeveel geschiedenis er in jullie plek zit.",
    "Familiebedrijf": "Tijdens het bekijken van jullie website viel me op dat het een echt familiebedrijf is.",
    "Duurzaam / B Corp": "Tijdens het bekijken van jullie website viel me op hoe bewust jullie met duurzaamheid omgaan.",
    "Net geopend / nieuw": "Tijdens het bekijken van jullie website viel me op dat jullie nog niet zo lang geleden zijn begonnen.",
    "Eigen verhaal/oprichter": "Tijdens het bekijken van jullie website viel me op hoeveel persoonlijk verhaal er achter jullie plek zit.",
    "Ambachtelijk / eigen productie": "Tijdens het bekijken van jullie website viel me op hoe ambachtelijk jullie te werk gaan.",
}

def _schone_citaatzin(text):
    """Zoek een korte, nette zin op de site die we LETTERLIJK kunnen citeren (altijd grammaticaal)."""
    text = re.sub(r"\s+", " ", text).strip()
    best, score = None, 0
    for z in re.split(r"(?<=[.!?]) ", text):
        z = z.strip()
        zl = z.lower()
        # streng: begint met hoofdletter, nette lengte, geen rare tekens of webrommel
        if not (35 <= len(z) <= 140): continue
        if not z[0].isupper(): continue
        if any(x in zl for x in ("cookie", "javascript", "©", "http", "@", "lees meer", "klik")): continue
        if any(ch in z for ch in "{}<>|[]"): continue
        s = sum(k in zl for k in KEYWORDS)
        if s > score: best, score = z.rstrip(".!?"), s
    return best if score >= 1 else None

def maak_opener(sigtext, seg, plaats, haak):
    """Altijd grammaticaal correcte openingszin: haak > letterlijk citaat > standaard."""
    if haak and haak in HOOK_OPENERS:
        return HOOK_OPENERS[haak]
    zin = _schone_citaatzin(sigtext or "")
    if zin:
        return f'Tijdens het bekijken van jullie website bleef deze zin me bij: "{zin}."'
    if plaats:
        return DEFAULT_OPENER[seg].rstrip(".") + f", hier in {plaats}."
    return DEFAULT_OPENER[seg]

def enrich_one(l):
    """Verrijk een enkele lead: e-mail, openingszin en social links. Muteert en geeft terug."""
    base = norm(l["website"])
    if not base:
        l["opener"] = l["opener"] or DEFAULT_OPENER[l["seg"]]
        return l
    dom = urlparse(base).netloc.replace("www.", "")
    email = l["email"]
    sigtext = ""  # verzamelde tekst van home + over-ons voor de haak-vinder en de openingszin
    for p in PATHS:
        url = base if p == "" else f"{base}/{p}"
        if not _robots_ok(base, p):
            continue
        try:
            r = polite_get(url, dom, timeout=8)
            if not r or r.status_code != 200 or "html" not in r.headers.get("content-type", ""): continue
            html = r.text
            if not email:
                c = re.findall(r"mailto:([^\"'?\s>]+)", html) + EMAIL_RE.findall(html)
                c = [x.strip().strip(".").lower() for x in c]
                c = [x for x in c if EMAIL_RE.fullmatch(x) and not any(b in x for b in BAD)]
                same = [x for x in c if x.split("@")[-1] == dom]; pool = same or c
                for pref in ("info@", "contact@", "welkom@", "boeking@"):
                    hit = next((x for x in pool if x.startswith(pref)), "")
                    if hit: email = hit; break
                if not email and pool: email = pool[0]
            for netwerk, rx in SOCIAL_RE.items():
                if not l.get(netwerk):
                    m = rx.search(html)
                    if m: l[netwerk] = m.group(0)
            if p in ("", "over-ons", "about"):
                sigtext += " " + detag(html)
            if email and sigtext: break
        except Exception:
            continue
    haak, heeft_app = detect_signals(sigtext)
    l["haak"] = haak
    l["heeft_app"] = heeft_app
    l["email"] = email
    if email and l["bron"] == "OpenStreetMap": l["bron"] = "OSM + website"
    l["opener"] = maak_opener(sigtext, l["seg"], l["plaats"], haak)
    return l

def has_mx(e):
    if not HAVE_DNS or "@" not in e: return None
    try: return len(dns.resolver.resolve(e.split("@")[-1], "MX")) > 0
    except Exception: return False

def voornaam(email):
    if not email: return ""
    lok = email.split("@")[0]
    if lok in ROLE or not re.fullmatch(r"[a-zA-Z]{3,14}", lok): return ""
    return lok.capitalize()

def ingang(l):
    e = l["email"]
    if not e: return "Adres ontbreekt"
    lok = e.split("@")[0]
    pers = lok not in ROLE and not any(lok.startswith(r) for r in ROLE)
    keten = any(k in (l["website"] + l["naam"]).lower() for k in KETEN)
    return f"{'Persoonlijk' if pers else 'Algemene inbox'} \u2022 {'keten (traag)' if keten else 'klein (snel)'}"

def maak_mail(l, brand):
    seg = l["seg"] if l["seg"] in STD else "natuurcamping"
    subject = SUBJECTS_STD.get(seg, SUBJECTS_STD["natuurcamping"])
    if l.get("heeft_app"):
        subject = "Geen tweede app \u2014 juist minder gedoe"
    # Aanhef: 'algemeen' = "Beste team," ; anders "Beste team van <naam>," (of voornaam indien bekend)
    if brand.get("aanhef_modus") == "algemeen":
        aanhef = "team"
    else:
        aanhef = l.get("_vn") or f"team van {l['naam']}"
    # Volledige-mail-override: vervangt de hele mail (alle segmenten).
    if brand.get("eigen_body", "").strip():
        ctx = {"aanhef": aanhef, "naam": l["naam"], "plaats": l["plaats"] or "jullie omgeving",
               "haak": l.get("haak", ""), "merk": brand["merk"], "links": brand["links"]}
        try:
            body = brand["eigen_body"].format(**ctx)
        except Exception:
            body = brand["eigen_body"]
        fu2 = FOLLOWUP_2.format(aanhef=aanhef, merk=brand["merk"])
        fu3 = FOLLOWUP_3.format(aanhef=aanhef)
        return subject, body, fu2, fu3
    if brand["eigen_mid"].strip():
        try:
            kern = brand["eigen_mid"].format(naam=l["naam"], plaats=l["plaats"] or "jullie omgeving",
                                             haak=l.get("haak", ""))
        except Exception:
            kern = brand["eigen_mid"]
    else:
        kern = STD[seg]["kern"]
    kern = kern.replace("VoiceStamp", brand["merk"])
    if l.get("heeft_app"):
        kern += "\n\n" + APP_LINE.replace("VoiceStamp", brand["merk"])
    linksblok = "Benieuwd hoe dat eruitziet? Neem gerust een kijkje:\n" + brand["links"]
    body = (f"Beste {aanhef},\n\n{kern}\n\n{linksblok}\n\n"
            f"Ik laat het graag zien als je denkt dat {brand['merk']} bij jullie past.\n\n"
            "Met vriendelijke groet,")
    fu2 = FOLLOWUP_2.format(aanhef=aanhef, merk=brand["merk"])
    fu3 = FOLLOWUP_3.format(aanhef=aanhef)
    return subject, body, fu2, fu3

def score_lead(l):
    s = 0
    if l["email"]: s += 40
    if l.get("mx") is True: s += 10
    ing = l.get("_ingang", "")
    if "Persoonlijk" in ing: s += 20
    if "klein" in ing: s += 15
    if l.get("haak"): s += 15
    if l.get("instagram") or l.get("linkedin"): s += 5
    return min(s, 100)

def kanaal_advies(l):
    if not l["email"]:
        if l.get("linkedin") or l.get("instagram"):
            return "Geen e-mail \u2014 benader via social"
        return "Geen e-mail \u2014 bel of zoek contact"
    ing = l.get("_ingang", "")
    if "keten" in ing:
        return "Keten \u2014 zoek de marketing/beslisser via LinkedIn"
    if "Persoonlijk" in ing:
        return "Mail de eigenaar direct (persoonlijk adres)"
    return "Mail (algemene inbox) \u2014 vraag om doorzetten"

def laad_verstuurd(bytes_data, filename):
    s = set()
    if not bytes_data: return s
    try:
        if filename.lower().endswith((".xlsx", ".xlsm")):
            wb = load_workbook(io.BytesIO(bytes_data), read_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for c in row:
                        if isinstance(c, str) and EMAIL_RE.fullmatch(c.strip().lower()):
                            s.add(c.strip().lower())
        else:
            for m in EMAIL_RE.findall(bytes_data.decode("utf-8", "ignore")): s.add(m.strip().lower())
    except Exception: pass
    return s

# ------------------------------------------------------------------ export
def build_xlsx(leads):
    wb = Workbook(); ws = wb.active; ws.title = "Leads"
    H = ["Score", "Naam", "Type", "Plaats", "E-mail", "Voornaam", "Haak (reden voor contact)",
         "Heeft app?", "Beste ingang", "Kanaal-advies", "E-mail geldig?", "Verzendbatch",
         "Website", "Telefoon", "Instagram", "LinkedIn", "Facebook", "Concept onderwerp",
         "Concept mail (1)", "Opvolgmail (2)", "Opvolgmail (3)", "Status", "Datum verstuurd", "Bron"]
    hf = PatternFill("solid", fgColor="2E5D4B"); hfont = Font("Arial", bold=True, color="FFFFFF")
    th = Side(style="thin", color="D0D0D0"); bd = Border(th, th, th, th)
    for c, h in enumerate(H, 1):
        x = ws.cell(1, c, h); x.fill = hf; x.font = hfont
        x.alignment = Alignment(vertical="center", wrap_text=True); x.border = bd
    for r, l in enumerate(leads, 2):
        mx = l.get("mx"); mxt = "ja" if mx is True else ("nee" if mx is False else "onbekend")
        vals = [l.get("_score", ""), l["naam"], l["type"], l["plaats"], l["email"], l.get("_vn", ""),
                l.get("haak", ""), "ja" if l.get("heeft_app") else "", l["_ingang"], l.get("_kanaal", ""),
                mxt, l.get("_batch", ""), l["website"], l["telefoon"], l.get("instagram", ""),
                l.get("linkedin", ""), l.get("facebook", ""), l["_subject"], l["_body"],
                l.get("_fu2", ""), l.get("_fu3", ""), "", "", l["bron"]]
        for c, v in enumerate(vals, 1):
            x = ws.cell(r, c, v); x.font = Font("Arial", size=10)
            x.alignment = Alignment(vertical="top", wrap_text=True); x.border = bd
    for i, w in enumerate([7, 24, 16, 13, 25, 10, 21, 9, 22, 30, 11, 10, 22, 13, 19, 19, 19, 28, 54, 40, 40, 10, 12, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 28
    ws.auto_filter.ref = f"A1:{get_column_letter(len(H))}{len(leads)+1}"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def build_csv(leads):
    cols = ["score", "email", "first_name", "company", "city", "hook", "has_app", "channel_advice",
            "custom_opener", "custom_subject", "custom_body", "followup_2", "followup_3",
            "send_batch", "website", "phone", "instagram", "linkedin", "facebook", "beste_ingang"]
    buf = io.StringIO(); w = csv.DictWriter(buf, fieldnames=cols); w.writeheader()
    for l in leads:
        if not l["email"]: continue
        w.writerow({"score": l.get("_score", ""), "email": l["email"], "first_name": l.get("_vn", ""),
                    "company": l["naam"], "city": l["plaats"], "hook": l.get("haak", ""),
                    "has_app": "yes" if l.get("heeft_app") else "", "channel_advice": l.get("_kanaal", ""),
                    "custom_opener": l.get("opener", ""), "custom_subject": l["_subject"],
                    "custom_body": l["_body"], "followup_2": l.get("_fu2", ""),
                    "followup_3": l.get("_fu3", ""), "send_batch": l.get("_batch", ""),
                    "website": l["website"], "phone": l["telefoon"], "instagram": l.get("instagram", ""),
                    "linkedin": l.get("linkedin", ""), "facebook": l.get("facebook", ""),
                    "beste_ingang": l["_ingang"]})
    return buf.getvalue().encode("utf-8")

# ------------------------------------------------------------------ pipeline
def draai(preset, area, whole, scrape, mx_check, dedupe_dom, n_per_dag, max_sites,
          workers, verstuurd, brand, progress, status):
    status("OpenStreetMap ophalen ...")
    keys_items = tuple((k, tuple(v)) for k, v in preset["keys"].items())
    els = osm_cached(area, keys_items, whole)
    leads = parse(els, preset["typemap"])
    seen, uniek = set(), []
    for l in leads:
        k = (l["naam"].lower(), l["plaats"].lower())
        if k not in seen: seen.add(k); uniek.append(l)
    leads = uniek
    if verstuurd:
        leads = [l for l in leads if l["email"].lower() not in verstuurd]
    status(f"{len(leads)} plekken gevonden. Websites verrijken ...")
    progress(0.15)

    if scrape:
        todo = [l for l in leads if l["website"]]
        if max_sites: todo = todo[:max_sites]
        done = 0
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futs = {ex.submit(enrich_one, l): l for l in todo}
            for _ in as_completed(futs):
                done += 1
                if done % 5 == 0 or done == len(todo):
                    status(f"Websites verrijken … {done}/{len(todo)}")
                    progress(0.15 + 0.7 * done / max(len(todo), 1))
    progress(0.86)

    if dedupe_dom:
        gz, dd = set(), []
        for l in leads:
            d = (l["email"].split("@")[-1] if l["email"] else l["website"]).lower()
            if d and d in gz: continue
            gz.add(d); dd.append(l)
        leads = dd

    for l in leads:
        l["_vn"] = voornaam(l["email"]); l["_ingang"] = ingang(l)

    if mx_check and HAVE_DNS:
        status("E-mailadressen controleren (MX) ...")
        with ThreadPoolExecutor(max_workers=workers) as ex:
            res = {ex.submit(has_mx, l["email"]): l for l in leads if l["email"]}
            for fut in as_completed(res):
                res[fut]["mx"] = fut.result()
        for l in leads:
            if "mx" not in l: l["mx"] = None
    else:
        for l in leads: l["mx"] = None
    progress(0.93)

    for l in leads:
        l["_score"] = score_lead(l)
        l["_kanaal"] = kanaal_advies(l)
        l["_subject"], l["_body"], l["_fu2"], l["_fu3"] = maak_mail(l, brand)

    leads.sort(key=lambda l: (-l["_score"], l["type"], l["naam"]))
    t = 0
    for l in leads:
        if l["email"] and l.get("mx") is not False:
            l["_batch"] = f"Dag {t // max(n_per_dag, 1) + 1}"; t += 1
        else: l["_batch"] = ""
    progress(1.0)
    return leads

# ------------------------------------------------------------------ UI
st.set_page_config(page_title="VoiceStamp Lead Finder", page_icon="🎙️", layout="wide")

# ---- Huisstijl (thema-veilig: vecht niet tegen light/dark mode) ----
ACCENT = "#2E5D4B"   # VoiceStamp-groen; ook in .streamlit/config.toml zetten
st.markdown(f"""
<style>
  .vs-header {{ display:flex; align-items:center; gap:14px; margin: 2px 0 6px 0; }}
  .vs-logo {{ width:46px; height:46px; border-radius:12px; background:{ACCENT}; color:#fff;
             display:flex; align-items:center; justify-content:center; font-size:24px; flex:0 0 auto; }}
  .vs-title {{ font-size:30px; font-weight:800; line-height:1.05; }}
  .vs-sub {{ opacity:0.7; margin-top:3px; font-size:14px; }}
  .stButton>button {{ border-radius:10px; font-weight:600; }}
  .stTabs [data-baseweb="tab-list"] {{ gap:4px; }}
  .stTabs [data-baseweb="tab"] {{ border-radius:10px 10px 0 0; padding:8px 16px; }}
  .vs-step {{ padding:14px 16px; border:1px solid rgba(128,128,128,0.25); border-radius:14px;
             height:100%; }}
  .vs-step b {{ color:{ACCENT}; }}
</style>
""", unsafe_allow_html=True)

st.markdown(
    '<div class="vs-header"><div class="vs-logo">🎙️</div>'
    '<div><div class="vs-title">VoiceStamp Lead Finder</div>'
    '<div class="vs-sub">Vind leads, schrijf de mails, beheer je opvolging.</div></div></div>',
    unsafe_allow_html=True)
st.write("")

with st.sidebar:
    st.header("Instellingen")
    campagne = st.selectbox("Campagne", list(PRESETS.keys()))
    heel = st.checkbox("Heel Nederland (traag, tabblad open houden)", value=False)
    provincie = st.selectbox("Provincie", PROVINCIES, index=3, disabled=heel)
    st.divider()
    scrape = st.checkbox("Websites verrijken (e-mail, socials, openingszin)", value=True)
    mx_check = st.checkbox("E-mail controleren (MX)", value=True, disabled=not HAVE_DNS)
    dedupe_dom = st.checkbox("Max 1 per domein/keten", value=False)
    snelheid = st.select_slider("Snelheid (parallelle verwerking)",
                                options=[4, 6, 8, 10, 12], value=8,
                                help="Hoger = sneller. De app blijft altijd netjes tegenover elk "
                                     "afzonderlijk domein (minimaal 1 sec ertussen).")
    n_per_dag = st.number_input("Aantal per verzenddag", 5, 200, 25, step=5)
    max_sites = st.number_input("Max sites verrijken (0 = alle)", 0, 5000, 0, step=10)
    st.divider()
    up = st.file_uploader("Al-verstuurde lijst (csv/xlsx, optioneel)", type=["csv", "xlsx", "xlsm"])
    with st.expander("Jouw merk & pitch (optioneel)"):
        merk = st.text_input("Productnaam", "VoiceStamp")
        aanhef_keuze = st.radio(
            "Aanhef", ["Beste team van [bedrijfsnaam]", "Beste team"], index=0,
            help="Persoonlijk (met bedrijfsnaam) scoort meestal beter. 'Beste team' is voor "
                 "iedereen identiek.")
        links_blok = st.text_area("Links-blok in de mail", LINKS, height=110)
        eigen_mid = st.text_area(
            "Eigen pitch (optioneel)", "",
            help="Vervangt de standaard segmenttekst (het middenstuk) voor de hele campagne. "
                 "Gebruik {naam}, {plaats}, {haak}. Laat leeg voor de ingebouwde teksten.")
        eigen_body = st.text_area(
            "Volledige standaardmail (optioneel)", "", height=200,
            help="Vervangt de HELE mail (opening + midden + slot) voor alle segmenten. Gebruik "
                 "{aanhef}, {naam}, {plaats}, {haak}, {merk}, {links}. Laat leeg om de ingebouwde "
                 "mails per segment te gebruiken. Eindig zelf met 'Met vriendelijke groet,'.")
    start = st.button("🚀 Start", type="primary", use_container_width=True)
    st.caption("Houd dit tabblad open tijdens het zoeken. Werk per provincie voor een snelle run. "
               "Dezelfde provincie opnieuw ophalen gaat direct (opgeslagen).")

    st.divider()
    with st.expander("Account / CRM", expanded=False):
        if not crm_beschikbaar():
            st.caption("CRM nog niet ingesteld. Zie LEES_MIJ_crm.md om je gratis database te koppelen.")
        elif st.session_state.get("sb_user"):
            st.success(f"Ingelogd als {st.session_state.get('sb_email','')}")
            if st.button("Uitloggen"):
                for k in ("sb_client", "sb_user", "sb_email",
                          "zoho_user", "zoho_pw", "zoho_host", "zoho_from", "zoho_sig"):
                    st.session_state.pop(k, None)
                st.rerun()
        else:
            st.text_input("E-mail", key="login_email")
            st.text_input("Wachtwoord", type="password", key="login_pw")
            cA, cB = st.columns(2)
            _em = (st.session_state.get("login_email") or "").strip()
            _pw = st.session_state.get("login_pw") or ""
            if cA.button("Inloggen"):
                if not _em or not _pw:
                    st.error("Vul een e-mail en wachtwoord in.")
                else:
                    try:
                        c, u = crm_auth(_em, _pw, registreren=False)
                        if u:
                            st.session_state["sb_client"] = c
                            st.session_state["sb_user"] = u.id
                            st.session_state["sb_email"] = _em
                            st.rerun()
                        else:
                            st.error("Inloggen mislukt. Controleer je gegevens.")
                    except Exception as e:
                        st.error(f"Inloggen mislukt: {e}")
            if cB.button("Registreren"):
                if not _em or not _pw:
                    st.error("Vul een e-mail en wachtwoord in.")
                elif len(_pw) < 6:
                    st.error("Wachtwoord moet minstens 6 tekens zijn.")
                else:
                    try:
                        crm_auth(_em, _pw, registreren=True)
                        st.info("Account aangemaakt. Bevestig eventueel je e-mail en log daarna in.")
                    except Exception as e:
                        st.error(f"Registreren mislukt: {e}")

    with st.expander("Mailen (Zoho)", expanded=False):
        if not st.session_state.get("sb_user"):
            st.caption("Log eerst in bij **Account / CRM** hierboven. De mailfunctie is alleen "
                       "beschikbaar als je bent ingelogd.")
        elif st.session_state.get("zoho_user"):
            st.success(f"Mail actief: {st.session_state['zoho_user']}")
            rest = MAIL_DAG_LIMIET
            if st.session_state.get("sb_user"):
                rest = MAIL_DAG_LIMIET - crm_sent_today(st.session_state["sb_client"], st.session_state["sb_user"])
            else:
                rest = MAIL_DAG_LIMIET - st.session_state.get("zoho_sent_today", 0)
            st.caption(f"Nog {max(rest,0)} van {MAIL_DAG_LIMIET} mails vandaag.")
            if st.button("Mail uitloggen"):
                for k in ("zoho_user", "zoho_pw", "zoho_host", "zoho_from", "zoho_sig"):
                    st.session_state.pop(k, None)
                st.rerun()
        else:
            st.caption("Log in om rechtstreeks vanuit de app te mailen (max 10 per dag).")
            zu = st.text_input("Zoho e-mail", key="zoho_email_in")
            zp = st.text_input("App-wachtwoord", type="password", key="zoho_pw_in",
                               help="Maak in Zoho een app-specifiek wachtwoord aan (niet je gewone wachtwoord).")
            zh = st.selectbox("Server", ["smtp.zoho.eu", "smtp.zoho.com"], key="zoho_host_in")
            zf = st.text_input("Afzendernaam", "Yusuf Tatlicioglu", key="zoho_from_in")
            zsig = st.text_area("Handtekening (onder de mail)",
                                "Yusuf Tatlicioglu\n0646756497\n\nFounder VoiceStamp\nhttps://www.voicestamp.nl/",
                                key="zoho_sig_in", height=90,
                                help="Wordt onder 'Met vriendelijke groet,' geplaatst bij app-verzending.")
            if st.button("Mail inloggen"):
                if not (zu.strip() and zp):
                    st.error("Vul je Zoho-e-mail en app-wachtwoord in.")
                else:
                    st.session_state["zoho_user"] = zu.strip()
                    st.session_state["zoho_pw"] = zp
                    st.session_state["zoho_host"] = zh
                    st.session_state["zoho_from"] = zf.strip() or zu.strip()
                    st.session_state["zoho_sig"] = zsig
                    st.rerun()

brand = {"merk": merk.strip() or "VoiceStamp", "links": links_blok, "eigen_mid": eigen_mid,
         "eigen_body": eigen_body,
         "aanhef_modus": "algemeen" if aanhef_keuze == "Beste team" else "persoonlijk"}

if start:
    preset = PRESETS[campagne]
    verstuurd = laad_verstuurd(up.getvalue(), up.name) if up else set()
    bar = st.progress(0.0); box = st.empty(); t0 = time.time()
    try:
        leads = draai(preset, None if heel else provincie, heel, scrape, mx_check,
                      dedupe_dom, n_per_dag, max_sites, int(snelheid), verstuurd, brand,
                      lambda p: bar.progress(min(p, 1.0)), lambda m: box.info(m))
        st.session_state["leads"] = leads
        box.success(f"Klaar in {int(time.time()-t0)} seconden.")
    except Exception as e:
        box.error(f"Er ging iets mis: {e}")

tab_res, tab_crm = st.tabs(["📋 Resultaten & mails", "📇 Mijn CRM"])

with tab_res:
    if st.session_state.get("leads"):
        leads = st.session_state["leads"]
        met = sum(1 for l in leads if l["email"])
        haken = sum(1 for l in leads if l.get("haak"))
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Leads totaal", len(leads))
        c2.metric("Met e-mail", met)
        c3.metric("Met haak", haken)
        c4.metric("Heeft app", sum(1 for l in leads if l.get("heeft_app")))
        c5.metric("Verzenddagen", (met // max(int(n_per_dag), 1)) + 1 if met else 0)

        f1, f2, f3 = st.columns([2, 1, 1])
        min_score = f1.slider("Toon alleen leads met score \u2265", 0, 100, 0, step=5)
        alleen_mail = f2.checkbox("Alleen met e-mail", value=False)
        alleen_haak = f3.checkbox("Alleen met haak", value=False)
        zicht = [l for l in leads if l.get("_score", 0) >= min_score
                 and (l["email"] or not alleen_mail) and (l.get("haak") or not alleen_haak)]

        df = pd.DataFrame([{"Score": l.get("_score", 0), "Naam": l["naam"], "Type": l["type"],
                            "Plaats": l["plaats"], "E-mail": l["email"], "Haak": l.get("haak", ""),
                            "App?": "ja" if l.get("heeft_app") else "", "Kanaal": l.get("_kanaal", ""),
                            "Batch": l.get("_batch", "")} for l in zicht])
        st.caption(f"{len(zicht)} van {len(leads)} leads getoond, gesorteerd op score.")
        st.dataframe(df, use_container_width=True, height=380, hide_index=True,
                     column_config={"Score": st.column_config.ProgressColumn(
                         "Score", min_value=0, max_value=100, format="%d")})

        # --- Kopieer-en-plak: volledige mail per lead ---
        st.subheader("Mail kopiëren per lead")
        mailbaar = [l for l in zicht if l["email"]]
        if mailbaar:
            keuze = st.selectbox(
                "Kies een lead",
                range(len(mailbaar)),
                format_func=lambda i: f"{mailbaar[i]['naam']} — {mailbaar[i]['email']}"
                                      f" (score {mailbaar[i].get('_score', 0)})")
            l = mailbaar[keuze]
            info = f"**Aan:** {l['email']}"
            if l.get("_kanaal"): info += f"  •  **Advies:** {l['_kanaal']}"
            if l.get("haak"): info += f"  •  **Haak:** {l['haak']}"
            st.markdown(info)
            tab1, tab2, tab3 = st.tabs(["Mail 1", "Opvolgmail 2", "Opvolgmail 3"])
            with tab1:
                onderwerp = st.text_input("Onderwerp", l["_subject"], key=f"subj_{keuze}")
                mailtekst = st.text_area("Mailtekst (pas gerust aan voor deze lead)",
                                         l["_body"], height=320, key=f"body_{keuze}")
                if st.session_state.get("sb_user") and st.session_state.get("zoho_user"):
                    # dagteller
                    gedaan = crm_sent_today(st.session_state["sb_client"], st.session_state["sb_user"])
                    rest = MAIL_DAG_LIMIET - gedaan
                    if st.button(f"✉️ Verstuur via Zoho ({max(rest,0)} over vandaag)", key=f"send_{keuze}"):
                        if rest <= 0:
                            st.warning("Je dagelijkse 10 mails zijn verstuurd. Morgen weer.")
                        else:
                            body_send = mailtekst + "\n" + st.session_state.get("zoho_sig", "")
                            try:
                                zoho_send(st.session_state["zoho_host"], 465,
                                          st.session_state["zoho_user"], st.session_state["zoho_pw"],
                                          st.session_state.get("zoho_from", st.session_state["zoho_user"]),
                                          l["email"], onderwerp, body_send)
                                st.success(f"Mail verstuurd naar {l['email']}.")
                                st.session_state["zoho_sent_today"] = gedaan + 1
                                if st.session_state.get("sb_user"):
                                    from datetime import date as _d
                                    cl = st.session_state["sb_client"]; uid = st.session_state["sb_user"]
                                    lid = crm_lead_id_by_email(cl, uid, l["email"])
                                    if not lid:
                                        crm_save(cl, uid, [l])
                                        lid = crm_lead_id_by_email(cl, uid, l["email"])
                                    if lid:
                                        crm_update(cl, lid, uid,
                                                   {"status": "Gemaild", "last_mailed": _d.today().isoformat()})
                                        crm_add_activity(cl, uid, lid, "Mail verstuurd", onderwerp)
                            except Exception as e:
                                st.error(f"Versturen mislukt: {e}")
                elif not st.session_state.get("sb_user"):
                    st.caption("Wil je direct vanuit de app mailen? Log links in bij **Account / CRM** "
                               "en daarna bij **Mailen (Zoho)**.")
                else:
                    st.caption("Wil je direct vanuit de app mailen? Log links in bij **Mailen (Zoho)**.")
            with tab2:
                st.code(l.get("_fu2", ""), language=None)
            with tab3:
                st.code(l.get("_fu3", ""), language=None)
            st.caption("Tip: klik rechtsboven in het grijze vak op het kopieer-icoon om de mail te kopiëren.")

            # --- Alle mails van één verzenddag in één keer ---
            st.markdown("**Of pak een hele verzenddag in één keer:**")
            dagen = sorted({l["_batch"] for l in mailbaar if l.get("_batch")},
                           key=lambda d: int(d.split()[-1]) if d.split()[-1].isdigit() else 999)
            if dagen:
                dag = st.selectbox("Verzenddag", dagen)
                dagleads = [l for l in mailbaar if l.get("_batch") == dag]
                blok = "\n\n════════════════════\n\n".join(
                    f"AAN: {l['email']}\nONDERWERP: {l['_subject']}\n\n{l['_body']}" for l in dagleads)
                st.caption(f"{len(dagleads)} mails in {dag}.")
                st.code(blok, language=None)
                st.download_button(f"⬇️ Excel van {dag}", build_xlsx(dagleads),
                                   file_name=f"voicestamp_leads_{dag.replace(' ', '_').lower()}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Geen leads met e-mail in de huidige selectie.")

        st.markdown("**Download de volledige selectie:**")
        d1, d2 = st.columns(2)
        d1.download_button("⬇️ Excel (alle kolommen: mails, haak, score, kanaal)", build_xlsx(zicht),
                           file_name="voicestamp_leads.xlsx", use_container_width=True,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        d2.download_button("⬇️ CSV voor Smartlead/Instantly", build_csv(zicht),
                           file_name="voicestamp_leads.csv", mime="text/csv", use_container_width=True)

        if st.session_state.get("sb_user"):
            if st.button("💾 Deze selectie opslaan in mijn CRM", use_container_width=True):
                try:
                    n = crm_save(st.session_state["sb_client"], st.session_state["sb_user"], zicht)
                    # log 'Lead opgeslagen' voor de nieuw toegevoegde leads
                    try:
                        cl = st.session_state["sb_client"]; uid = st.session_state["sb_user"]
                        for l in zicht:
                            if l.get("email"):
                                lid = crm_lead_id_by_email(cl, uid, l["email"])
                                if lid:
                                    acts = crm_load_activities(cl, uid, lid)
                                    if not acts:
                                        crm_add_activity(cl, uid, lid, "Lead opgeslagen", l.get("naam", ""))
                    except Exception:
                        pass
                    st.success(f"{n} nieuwe leads opgeslagen in je CRM (dubbele e-mails overgeslagen).")
                except Exception as e:
                    st.error(f"Opslaan mislukt: {e}")
        elif crm_beschikbaar():
            st.caption("Log links in bij Account / CRM om deze leads op te slaan.")
    else:
        st.markdown("#### Zo werkt het")
        s1, s2, s3 = st.columns(3)
        s1.markdown('<div class="vs-step"><b>1. Kies</b><br>Selecteer links een campagne '
                    'en provincie en klik op Start.</div>', unsafe_allow_html=True)
        s2.markdown('<div class="vs-step"><b>2. Bekijk</b><br>Je krijgt leads met e-mail, een haak, '
                    'een score en een kant-en-klare mail per lead.</div>', unsafe_allow_html=True)
        s3.markdown('<div class="vs-step"><b>3. Werk bij</b><br>Kopieer de mails, download alles, '
                    'of sla op in je CRM en volg later op.</div>', unsafe_allow_html=True)
        st.write("")
        st.info("Klaar om te beginnen? Kies links een campagne en provincie en klik op **Start**.")

with tab_crm:
    if not st.session_state.get("sb_user"):
        if crm_beschikbaar():
            st.info("Log links in bij **Account / CRM** om je opgeslagen leads te beheren.")
        else:
            st.info("Het CRM is nog niet ingesteld. Zie **LEES_MIJ_crm.md** om je gratis database te koppelen.")
    else:
            st.divider()
            st.header("📇 Mijn CRM")

            with st.expander("📥 Bestaande Excel importeren"):
                st.caption("Zet je bestaande leadlijst (bijv. VoiceStamp_leads.xlsx) in \u00e9\u00e9n keer "
                           "in je CRM. Dubbele e-mailadressen worden overgeslagen. Bestaande statussen "
                           "blijven behouden.")
                imp = st.file_uploader("Kies je Excel of CSV", type=["xlsx", "xlsm", "csv"],
                                       key="crm_import_file")
                if imp is not None and st.button("Importeren", key="crm_import_btn"):
                    try:
                        toe, over = crm_import_uit_bytes(
                            st.session_state["sb_client"], st.session_state["sb_user"],
                            imp.getvalue(), imp.name)
                        st.success(f"{toe} leads ge\u00efmporteerd, {over} overgeslagen (dubbel of geen e-mail).")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Import mislukt: {e}")

            try:
                data = crm_load(st.session_state["sb_client"], st.session_state["sb_user"])
            except Exception as e:
                data = []
                st.error(f"Kon CRM niet laden: {e}")
            if not data:
                st.caption("Nog geen leads opgeslagen. Zoek hierboven en klik op 'Opslaan in mijn CRM'.")
            else:
                from datetime import date as _date
                cdf = pd.DataFrame(data)
                for kol, leeg in [("note", ""), ("status", "Nieuw"), ("segment", ""),
                                  ("company_name", ""), ("email", ""), ("score", 0),
                                  ("next_action", ""), ("next_action_date", None), ("dnc_reason", ""),
                                  ("last_mailed", None)]:
                    if kol not in cdf.columns:
                        cdf[kol] = leeg
                cdf["note"] = cdf["note"].fillna("")
                cdf["next_action"] = cdf["next_action"].fillna("")
                cdf["dnc_reason"] = cdf["dnc_reason"].fillna("")

                # ---- Dashboard ----
                vandaag = _date.today().isoformat()
                def _telt(s): return int((cdf["status"] == s).sum())
                actief = ~cdf["status"].isin(["Klant", "Afgewezen", "Do not contact"])
                opvolg = cdf[(cdf["next_action_date"].notna()) &
                             (cdf["next_action_date"].astype(str) <= vandaag) & actief]
                st.markdown("### Mijn overzicht")
                m = st.columns(6)
                m[0].metric("Nieuw", _telt("Nieuw"))
                m[1].metric("Gemaild", _telt("Gemaild"))
                m[2].metric("Follow-up nodig", len(opvolg))
                m[3].metric("Reacties", _telt("Reactie"))
                m[4].metric("Afspraken", _telt("Afspraak"))
                m[5].metric("Klanten", _telt("Klant"))

                if len(opvolg):
                    st.markdown("#### 🔥 Vandaag opvolgen")
                    st.dataframe(
                        opvolg[["company_name", "status", "next_action", "next_action_date"]]
                        .rename(columns={"company_name": "Bedrijf", "status": "Status",
                                         "next_action": "Volgende actie", "next_action_date": "Datum"}),
                        use_container_width=True, hide_index=True)

                st.divider()
                st.markdown("### Alle leads")
                stat = st.multiselect("Filter op status", STATUS_OPTIES, default=[])
                toon = cdf[cdf["status"].isin(stat)] if stat else cdf
                st.caption(f"{len(toon)} van {len(cdf)} leads.")
                edit = st.data_editor(
                    toon[["company_name", "segment", "email", "score", "status",
                          "next_action", "next_action_date", "last_mailed", "note", "dnc_reason"]],
                    use_container_width=True, height=420, key="crm_editor",
                    column_config={
                        "company_name": st.column_config.TextColumn("Bedrijf", disabled=True),
                        "segment": st.column_config.TextColumn("Type", disabled=True),
                        "email": st.column_config.TextColumn("E-mail", disabled=True),
                        "score": st.column_config.ProgressColumn("Score", min_value=0, max_value=100, format="%d"),
                        "status": st.column_config.SelectboxColumn("Status", options=STATUS_OPTIES),
                        "next_action": st.column_config.TextColumn("Volgende actie"),
                        "next_action_date": st.column_config.DateColumn("Datum volg. actie"),
                        "last_mailed": st.column_config.DateColumn("Gemaild op", disabled=True),
                        "note": st.column_config.TextColumn("Notitie"),
                        "dnc_reason": st.column_config.TextColumn("Reden (Do not contact)"),
                    })
                if st.button("Wijzigingen opslaan", type="primary"):
                    fouten = 0
                    cl = st.session_state["sb_client"]; uid = st.session_state["sb_user"]
                    for idx in edit.index:
                        rid = cdf.loc[idx, "id"]
                        nd = edit.loc[idx, "next_action_date"]
                        nd = None if (nd is None or pd.isna(nd)) else str(nd)[:10]
                        nieuwe_status = edit.loc[idx, "status"]
                        velden = {"status": nieuwe_status,
                                  "next_action": edit.loc[idx, "next_action"] or "",
                                  "next_action_date": nd,
                                  "note": edit.loc[idx, "note"] or "",
                                  "dnc_reason": edit.loc[idx, "dnc_reason"] or ""}
                        try:
                            crm_update(cl, rid, uid, velden)
                            if nieuwe_status != cdf.loc[idx, "status"]:
                                crm_add_activity(cl, uid, rid, "Status gewijzigd", f"→ {nieuwe_status}")
                        except Exception:
                            fouten += 1
                    if fouten:
                        st.warning(f"Opgeslagen, maar {fouten} rijen gaven een fout.")
                    else:
                        st.success("Wijzigingen opgeslagen.")

                # ---- Verwijderen ----
                st.divider()
                st.markdown("### Leads verwijderen")
                cl = st.session_state["sb_client"]; uid = st.session_state["sb_user"]
                vd1, vd2 = st.columns(2)
                with vd1:
                    weg = st.multiselect(
                        "Kies leads om te verwijderen",
                        options=list(cdf.index),
                        format_func=lambda i: f"{cdf.loc[i,'company_name']} ({cdf.loc[i,'email']})")
                    if st.button("Verwijder gekozen leads") and weg:
                        fout = 0
                        for i in weg:
                            try:
                                crm_delete(cl, cdf.loc[i, "id"], uid)
                            except Exception:
                                fout += 1
                        if fout:
                            st.warning(f"{len(weg)-fout} verwijderd, {fout} met een fout.")
                        else:
                            st.success(f"{len(weg)} leads verwijderd.")
                        st.rerun()
                with vd2:
                    st.caption("Alles wissen kan niet ongedaan worden gemaakt.")
                    bevestig = st.checkbox("Ja, verwijder al mijn leads")
                    if st.button("🗑️ Alle leads verwijderen", type="secondary") and bevestig:
                        try:
                            n = crm_delete_all(cl, uid)
                            st.success(f"Alle leads verwijderd ({n}).")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Verwijderen mislukt: {e}")

                # ---- Mail versturen vanuit CRM (Zoho) ----
                st.divider()
                st.markdown("### Mail versturen")
                if not st.session_state.get("zoho_user"):
                    st.caption("Log links in bij **Mailen (Zoho)** om vanuit je CRM te mailen.")
                else:
                    mlijst = cdf[cdf["email"].fillna("") != ""]
                    if not len(mlijst):
                        st.caption("Geen leads met e-mail in je CRM.")
                    else:
                        idxs = list(mlijst.index)
                        mk = st.selectbox(
                            "Kies een lead om te mailen", idxs,
                            format_func=lambda i: f"{cdf.loc[i,'company_name']} ({cdf.loc[i,'email']})",
                            key="crm_mail_sel")
                        seg = LABEL_TO_SEG.get(str(cdf.loc[mk, "segment"]), "natuurcamping")
                        lead_obj = {"seg": seg, "naam": cdf.loc[mk, "company_name"] or "",
                                    "plaats": "", "_vn": "", "haak": "", "heeft_app": False,
                                    "email": cdf.loc[mk, "email"]}
                        subj, body, _, _ = maak_mail(lead_obj, brand)
                        ondw = st.text_input("Onderwerp", subj, key="crm_mail_subj")
                        mtekst = st.text_area("Mailtekst (pas gerust aan voor deze lead)",
                                              body, height=320, key="crm_mail_body")
                        gedaan = crm_sent_today(st.session_state["sb_client"], st.session_state["sb_user"])
                        rest = MAIL_DAG_LIMIET - gedaan
                        if st.button(f"✉️ Verstuur via Zoho ({max(rest,0)} over vandaag)", key="crm_send"):
                            if rest <= 0:
                                st.warning("Je dagelijkse 10 mails zijn verstuurd. Morgen weer.")
                            else:
                                body_send = mtekst + "\n" + st.session_state.get("zoho_sig", "")
                                try:
                                    from datetime import date as _d
                                    zoho_send(st.session_state["zoho_host"], 465,
                                              st.session_state["zoho_user"], st.session_state["zoho_pw"],
                                              st.session_state.get("zoho_from", st.session_state["zoho_user"]),
                                              lead_obj["email"], ondw, body_send)
                                    cl = st.session_state["sb_client"]; uid = st.session_state["sb_user"]
                                    crm_update(cl, cdf.loc[mk, "id"], uid,
                                               {"status": "Gemaild", "last_mailed": _d.today().isoformat()})
                                    crm_add_activity(cl, uid, cdf.loc[mk, "id"], "Mail verstuurd", ondw)
                                    st.success(f"Mail verstuurd naar {lead_obj['email']}.")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Versturen mislukt: {e}")

                # ---- Contacthistorie per lead ----
                st.divider()
                st.markdown("### Contacthistorie")
                namen = cdf["company_name"].tolist()
                if namen:
                    kz = st.selectbox("Kies een lead", range(len(namen)),
                                      format_func=lambda i: namen[i], key="hist_sel")
                    lid = cdf.loc[cdf.index[kz], "id"]
                    nieuwe_notitie = st.text_input("Notitie toevoegen", key="hist_note")
                    if st.button("Notitie loggen") and nieuwe_notitie.strip():
                        crm_add_activity(st.session_state["sb_client"], st.session_state["sb_user"],
                                         lid, "Notitie", nieuwe_notitie.strip())
                        st.success("Genoteerd.")
                        st.rerun()
                    acts = crm_load_activities(st.session_state["sb_client"],
                                               st.session_state["sb_user"], lid)
                    if acts:
                        for a in acts:
                            datum = str(a.get("created_at", ""))[:10]
                            oms = f" — {a['description']}" if a.get("description") else ""
                            st.markdown(f"- **{datum}**  {a.get('type','')}{oms}")
                    else:
                        st.caption("Nog geen activiteiten voor deze lead.")

st.divider()
st.caption(
    "Zakelijk gebruik. Benader alleen relevante bedrijven met een duidelijke afmeldmogelijkheid, "
    "en houd je aan de AVG en de Telecommunicatiewet (B2B-uitzondering). De data komt van "
    "OpenStreetMap (ODbL) en van openbare bedrijfswebsites. Verstuur gedoseerd; de verzendbatches "
    "helpen daarbij. Deze tool zoekt geen persoonsgegevens op en is bedoeld als hulpmiddel, niet "
    "als vervanging van je eigen beoordeling per lead.")
